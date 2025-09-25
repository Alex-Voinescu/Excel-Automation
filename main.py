import openpyxl as xl
from openpyxl.chart import BarChart, Reference, series


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    new_cell_name = sheet.cell(1, 4)
    new_cell_name.value = 'corrected_price'
    for row in range(2, sheet.max_row+1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value*0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=1,
                       max_row=sheet.max_row,
                       min_col=3,
                       max_col=4)

    labels = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=1,
                       max_col=4)

    chart = BarChart()
    chart.type = "col"
    chart.title = "Price vs Corrected Price"
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(labels)
    chart.legend.position = "b"
    sheet.add_chart(chart, 'f2')

    wb.save(filename)


process_workbook('transactions.xlsx')
